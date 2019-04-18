from openpyxl import load_workbook

import os,django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "WeChatPM.settings")        #  项目名称.settings
django.setup()

from backend.models import SuperUser


def xlsxHandler(localpath):
    wb = load_workbook(localpath)   # 打开已存在的execl文件，格式xlsx

    ws = wb[wb.sheetnames[0]]       # 选择第一张sheet表
    rows = ws.max_row               # 获取表的最大行数
    columns = ws.max_column         # 获取表的最大列数

    column_heading = [ws.cell(row=1,column=x).value for x in  range(1,columns+1)]   # 读取excel第一行的值，写入list

    column_name = ['用户名','公司名称','电话']      # 数据库必需字段

    # print(column_heading)         # 文件第一行title

    if len([name for name in column_name if name not in column_heading]) == 0:  # 返回字段组成的list为空，则说明文件列标题包含MySQL需要的字段
        print(' - 检查完成，执行写入')
        # 判断Excel中各字段所在列号
        username = column_heading.index(column_name[0])         # 用户名 - 位置
        company = column_heading.index(column_name[1])         # 公司 - 位置
        phone = column_heading.index(column_name[2])          # 电话 - 位置


        if ws.cell(row=2,column=1).value == None:
            table_start_line = 3
        else:
            table_start_line = 2
        data = []
        superuserList = []
        for row in range(table_start_line, rows + 1):
            # 此处用于筛选“三级分类”的关键字
            # if ws.cell(row=row, column=username+1).value not in List_Product:  # 判断是否为所需行
            #     pass
            # else:
            for column in range(1, columns + 1):  # 因为从第1列开始，所以此处从1开始
                data.append(str(ws.cell(row=row, column=column).value))  # 以字符串形式保存数据到MySQL
            print(data)
            print(data[username],data[company],data[phone])
            superuserList.append(SuperUser(username=data[username],company=data[company],phone=data[phone]))
            data = []
        print('superuserList',superuserList)

        try:
            SuperUser.objects.bulk_create(superuserList)
            msg = '导入成功'
        except Exception as e:
            print('异常',e)
            msg = '导入失败'
    else:
        print('文件列标题不完全包含数据库需要的字段，请检查文件。')
        msg = '文件列标题不完全包含数据库需要的字段，请检查文件。'
    wb.close() # 关闭excel

    return msg